import os
import glob
import requests
import time
import random
import openpyxl
from openpyxl import Workbook, load_workbook
import re
from pathlib import Path
from urllib.parse import urlparse

class MultiTypeProcessor:
    def __init__(self):

        #Data files
        self.data_folder = r"C:\Your\Data\Folder"
        self.main_txt = r"C:\Txt_file\Contains\Mixed_Data"
        
        # Excel files
        self.ip_excel = r"\ip_results.xlsx"
        self.domain_excel = r"\domain_results.xlsx"
        self.hash_excel = r"\hash_results.xlsx"
        self.url_excel = r"\url_results.xlsx"
        
        self.api_key = "YOUR_API_KEY"
        
        # VirusTotal API endpoints
        self.api_endpoints = {
            'ip': "https://www.virustotal.com/api/v3/ip_addresses/{data}",
            'domain': "https://www.virustotal.com/api/v3/domains/{data}",
            'hash': "https://www.virustotal.com/api/v3/files/{data}",
            'url': "https://www.virustotal.com/api/v3/urls/{data}"
        }
        
        self.session = requests.Session()
        self.session.headers.update({
            "HEADERS"
        })

    # Validation

    def is_valid_ip(self, ip: str) -> bool:
        pattern = r'^(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$'
        return re.match(pattern, ip.strip()) is not None

    def is_valid_domain(self, domain: str) -> bool:
        pattern = r'^(?:[a-zA-Z0-9](?:[a-zA-Z0-9-]{0,61}[a-zA-Z0-9])?\.)+[a-zA-Z]{2,}$'
        return re.match(pattern, domain.strip()) is not None

    def is_valid_hash(self, hash_str: str) -> bool:
        hash_str = hash_str.strip()
        # MD5 (32), SHA1 (40), SHA256 (64), SHA512 (128)
        if len(hash_str) in [32, 40, 64, 128] and re.match(r'^[a-fA-F0-9]+$', hash_str):
            return True
        return False

    def is_valid_url(self, url: str) -> bool:
        try:
            result = urlparse(url.strip())
            return all([result.scheme, result.netloc])
        except:
            return False

    def clean_and_classify_data(self, raw_data: str) -> tuple:

        # Cleaning Process

        cleaned = raw_data.strip()
        if cleaned.startswith('•'):
            cleaned = cleaned[1:].strip()
        
        # [.] -> . 
        cleaned = cleaned.replace('[.]', '.')
        
        # hxxps:// -> https:// 
        cleaned = cleaned.replace('hxxps://', 'https://')
        cleaned = cleaned.replace('hxxp://', 'http://')
        
        # Identify element
        if self.is_valid_ip(cleaned):
            return cleaned, 'ip'
        elif self.is_valid_hash(cleaned):
            return cleaned, 'hash'
        elif self.is_valid_url(cleaned):
            return cleaned, 'url'
        elif self.is_valid_domain(cleaned):
            return cleaned, 'domain'
        else:
            print(f"Unknown format: {raw_data}")
            return None, None

    def remove_duplicates_from_main_txt(self):
        """Cleaning duplicates"""

        if not os.path.exists(self.main_txt):
            return
            
        print("Checking for duplicates in main.txt...")
        
        with open(self.main_txt, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        unique_lines = []
        seen = set()
        duplicates_found = 0
        
        for line in lines:
            line_clean = line.strip()
            if line_clean and line_clean not in seen:
                seen.add(line_clean)
                unique_lines.append(line)
            elif line_clean in seen:
                duplicates_found += 1
        
        if duplicates_found > 0:
            with open(self.main_txt, 'w', encoding='utf-8') as f:
                f.writelines(unique_lines)
            print(f"{duplicates_found} duplicated lines removed.")
        else:
            print("Duplicated lines not found.")

    def merge_txt_files(self):
        """Merge and classify txt files"""
        print("TXT files merged...")
        
        # Remove duplicates
        self.remove_duplicates_from_main_txt()
        
        # Read current datas
        existing_data = {
            'ip': set(), 'domain': set(), 'hash': set(), 'url': set()
        }
        
        if os.path.exists(self.main_txt):
            with open(self.main_txt, 'r', encoding='utf-8') as f:
                for line in f:
                    cleaned_data, data_type = self.clean_and_classify_data(line)
                    if cleaned_data and data_type:
                        existing_data[data_type].add(cleaned_data)
        
        # Handle new files
        txt_files = glob.glob(os.path.join(self.data_folder, "*.txt"))
        print(f"{len(txt_files)} TXT files found.")
        
        new_data = {'ip': set(), 'domain': set(), 'hash': set(), 'url': set()}
        
        for txt_file in txt_files:
            print(f"Process: {os.path.basename(txt_file)}")
            try:
                with open(txt_file, 'r', encoding='utf-8') as f:
                    for line in f:
                        cleaned_data, data_type = self.clean_and_classify_data(line)
                        if cleaned_data and data_type:
                            if cleaned_data not in existing_data[data_type]:
                                new_data[data_type].add(cleaned_data)
            except Exception as e:
                print(f"Unexpected Error {txt_file}: {e}")
        
        # Add new data to main.txt
        if any(new_data.values()):
            with open(self.main_txt, 'a', encoding='utf-8') as f:
                for data_type in ['ip', 'domain', 'hash', 'url']:
                    for item in sorted(new_data[data_type]):
                        f.write(item + '\n')
            
            total_new = sum(len(v) for v in new_data.values())
            print(f"{total_new} new data added.")
            for data_type, items in new_data.items():
                if items:
                    print(f"  - {data_type}: {len(items)}")
        else:
            print("New data not found.")
        
        # Merge all data
        all_data = {}
        for data_type in ['ip', 'domain', 'hash', 'url']:
            all_data[data_type] = existing_data[data_type].union(new_data[data_type])
        
        return all_data

    def setup_excel(self, excel_file: str, data_type: str):
        """Prepare excel"""
        if not os.path.exists(excel_file):
            print(f"New {data_type} Excel file is being created...")
            wb = Workbook()
            ws = wb.active
            ws.title = f"{data_type.title()} Results"
            
            if data_type == 'ip':
                ws['A1'] = "IP Address"
            elif data_type == 'domain':
                ws['A1'] = "Domain"
            elif data_type == 'hash':
                ws['A1'] = "File Hash"
            elif data_type == 'url':
                ws['A1'] = "URL"
            
            ws['B1'] = "Threats (malicious)"
            ws['C1'] = "Total Engines"
            wb.save(excel_file)
            return wb, ws
        else:
            try:
                print(f"Current {data_type} Excel file is opening...")
                wb = load_workbook(excel_file)
                ws = wb.active
                return wb, ws
            except Exception as e:
                print(f"{data_type} Excel file cannot open: {e}")
                backup_name = excel_file.replace('.xlsx', '_corrupted_backup.xlsx')
                if os.path.exists(excel_file):
                    os.rename(excel_file, backup_name)
                    print(f"Corrupted file name: {backup_name}")
                return self.setup_excel(excel_file, data_type)

    def write_data_to_excel(self, all_data: dict):
        """Write data to excel"""
        print("Data are written to excel...")
        
        excel_files = {
            'ip': self.ip_excel,
            'domain': self.domain_excel,
            'hash': self.hash_excel,
            'url': self.url_excel
        }
        
        for data_type, data_set in all_data.items():
            if not data_set:
                continue
                
            excel_file = excel_files[data_type]
            
            try:
                wb, ws = self.setup_excel(excel_file, data_type)
                
                existing_data_in_excel = set()
                max_row = ws.max_row if ws.max_row > 1 else 1
                for row in range(2, max_row + 1):
                    val = ws[f'A{row}'].value
                    if val:
                        existing_data_in_excel.add(str(val).strip())
                
                new_data = sorted(data_set - existing_data_in_excel)
                if new_data:
                    next_row = ws.max_row + 1
                    print(f"{data_type}: {len(new_data)} new data adding...")
                    
                    for i, data_item in enumerate(new_data):
                        ws[f'A{next_row + i}'] = data_item
                        ws[f'B{next_row + i}'] = None
                        ws[f'C{next_row + i}'] = None
                        
                        if (i + 1) % 100 == 0:
                            print(f"  -> {i + 1}/{len(new_data)} {data_type} data added...")
                    
                    wb.save(excel_file)
                    print(f"✓ {data_type}: {len(new_data)} data added.")
                else:
                    print(f"{data_type}: new data not found.")
                    
                wb.close()
                
            except Exception as e:
                print(f"{data_type} Exception: {e}")
                raise

    def _vt_request(self, url: str, timeout=30):
        """VirusTotal API request"""
        headers = {"x-apikey": self.api_key, **self.session.headers}
        try:
            resp = self.session.get(url, headers=headers, timeout=timeout, verify=False)
            if resp.status_code == 429:
                print("  -> ⚠ Rate limit (429). Waiting 60 seconds...")
                time.sleep(60 + random.randint(0, 5))
                return self.session.get(url, headers=headers, timeout=timeout, verify=False)
            return resp
        except requests.exceptions.Timeout:
            print("  -> ✗ timeout")
            return None
        except requests.exceptions.ConnectionError as e:
            print(f"  -> ✗ Connection Error: {e}")
            return None
        except requests.exceptions.RequestException as e:
            print(f"  -> ✗ Request Exception: {e}")
            return None

    def get_virustotal_info(self, data: str, data_type: str):
        """Info from Virustotal"""
        print(f"  -> VirusTotal API query: {data} ({data_type})")
        
        # Base64 encoding
        if data_type == 'url':
            import base64
            url_id = base64.urlsafe_b64encode(data.encode()).decode().strip('=')
            url = self.api_endpoints[data_type].format(data=url_id)
        else:
            url = self.api_endpoints[data_type].format(data=data)
        
        resp = self._vt_request(url)
        if resp is None:
            return None

        if resp.status_code == 200:
            try:
                data_json = resp.json()
            except ValueError:
                print("  -> ✗ JSON parse error")
                return None

            attributes = (data_json or {}).get("data", {}).get("attributes", {})
            stats = attributes.get("last_analysis_stats", {}) or {}
            
            if not stats:
                print("  -> Not found: last_analysis_stats empty. Decided 0.")
                return (0, 0)

            malicious = int(stats.get("malicious", 0) or 0)
            suspicious = int(stats.get("suspicious", 0) or 0)
            harmless = int(stats.get("harmless", 0) or 0)
            timeout = int(stats.get("timeout", 0) or 0)
            undetected = int(stats.get("undetected", 0) or 0)

            total = malicious + suspicious + harmless + timeout + undetected
            print(f"  -> ✓ Success: {malicious}/{total} (malicious/total)")
            return (malicious, total)

        elif resp.status_code in (401, 403):
            print("  -> ✗ Authorization: API key is false or access denied (401/403)")
            return None
        elif resp.status_code == 404:
            print("  -> ✓ Record not found (404). Decided 0.")
            return (0, 0)
        else:
            print(f"  -> ✗ HTTP error: {resp.status_code}")
            return None

    def update_virustotal_info(self, data_type: str):
        """Update virustotal info"""
        excel_files = {
            'ip': self.ip_excel,
            'domain': self.domain_excel,
            'hash': self.hash_excel,
            'url': self.url_excel
        }
        
        excel_file = excel_files[data_type]
        if not os.path.exists(excel_file):
            print(f"{data_type} Excel file was not found.")
            return
            
        print(f"{data_type} VirusTotal API infos update...")
        WAIT_SECONDS = 20

        try:
            wb = load_workbook(excel_file)
            ws = wb.active
        except Exception as e:
            print(f"{data_type} Excel cannot open: {e}")
            return

        rows_to_process = []
        max_row = ws.max_row if ws.max_row > 1 else 1
        for row in range(2, max_row + 1):
            data_val = ws[f'A{row}'].value
            b_val = ws[f'B{row}'].value
            c_val = ws[f'C{row}'].value
            if data_val and (b_val in (None, "") or c_val in (None, "")):
                rows_to_process.append(row)

        total_items = len(rows_to_process)
        if total_items == 0:
            print(f"All {data_type} data has already been processed.")
            wb.close()
            return

        print(f"Number of {data_type}: {total_items}")
        print(f"Estimated total time: ~{(total_items * WAIT_SECONDS) // 60} minute\n")

        updated_count = 0
        for i, row in enumerate(rows_to_process, start=1):
            data_item = str(ws[f'A{row}'].value).strip()
            print(f"[{i}/{total_items}] Handling: {data_item} (line {row})")

            result = self.get_virustotal_info(data_item, data_type)
            if result is not None:
                malicious, total = result
                ws[f'B{row}'] = malicious
                ws[f'C{row}'] = total
                updated_count += 1

                if updated_count % 10 == 0:
                    try:
                        wb.save(excel_file)
                        print(f"  -> Excel saved ({updated_count} line updated)")
                    except Exception as e:
                        print(f"  -> Excel save error: {e}")

            if i < total_items:
                remaining = WAIT_SECONDS
                print(f"  -> Waiting {WAIT_SECONDS} second for rate limit...")
                while remaining > 0:
                    print(f"     Remaining: {remaining:02d} second", end='\r')
                    time.sleep(5)
                    remaining -= 5
                print("     Remaining: 00 sn      ")

        try:
            wb.save(excel_file)
            print(f"\n✓ {data_type} update done!")
            print(f"  - Updated {data_type}: {updated_count}")
        except Exception as e:
            print(f"Last save error: {e}")
        finally:
            wb.close()

    def run(self):
        print("Multi-Type process starting...\n")
        
        all_data = self.merge_txt_files()
        
        total_count = sum(len(v) for v in all_data.values())
        print(f"{total_count} unique data found.")
        for data_type, items in all_data.items():
            if items:
                print(f"  - {data_type}: {len(items)}")
        
        print()
        
        # Write to excel
        self.write_data_to_excel(all_data)
        print()
        
        # VirusTotal update
        for data_type in ['ip', 'domain', 'hash', 'url']:
            if all_data[data_type]:
                self.update_virustotal_info(data_type)
                print()
        
        print("DONE!")
        print(f"Main file: {self.main_txt}")
        print("Excel files:")
        print(f"  - IP: {self.ip_excel}")
        print(f"  - Domain: {self.domain_excel}")
        print(f"  - Hash: {self.hash_excel}")
        print(f"  - URL: {self.url_excel}")


def main():
    try:
        processor = MultiTypeProcessor()
        processor.run()
    except KeyboardInterrupt:
        print("\nProcess stopped by user.")
    except Exception as e:
        print(f"Unexpected error: {e}")
        input("Enter for continue...")

if __name__ == "__main__":
    main()